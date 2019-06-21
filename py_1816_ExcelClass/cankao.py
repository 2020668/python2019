# -*- coding: utf-8 -*-
"""

=================================
Author: keen
Created on: 2019/5/26

E-mail:keen2020@outlook.com

=================================


"""

import openpyxl


class Case:
    # 这个类用来存储用例
    pass


class ReadExcel_dict(object):
    """
    读取excel数据
    """

    def __init__(self, file_name, sheet_name, list1):
        """
        这个是用例初始化读取对象的
        :param file_name:  文件名字  -->  str
        :param sheet_name: 表单名字  -->  str
        :param list1: 指定列  -->  list
        """
        # 打开工作簿
        self.wb = openpyxl.load_workbook(file_name)
        # 选择表单
        self.sheet = self.wb[sheet_name]
        self.list = list1

    def read_data(self):
        """
        按行读取数据
        :return:  返回一个列表，列表中每个元素为一条用例
        """
        # 按行获取数据转换成列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)

        # 定义一个空列表用来存储所有的用例
        cases = []
        for i in range(2, self.sheet.max_row + 1):
            data = []
            data.append(self.sheet.cell(row=i, column=self.list[0]).value)  # 读取1列数据
            data.append(eval(self.sheet.cell(row=i, column=self.list[1]).value))  # 读取第3列数据
            print(data)


class ReadExcel_obj(object):
    """
    读取excel数据
    """

    def __init__(self, file_name, sheet_name, list1):
        """
        这个是用例初始化读取对象的
        :param file_name:  文件名字  -->  str
        :param sheet_name: 表单名字  -->  str
        """
        # 打开工作簿
        self.wb = openpyxl.load_workbook(file_name)
        # 选择表单
        self.sheet = self.wb[sheet_name]
        self.list = list1

    def read_data_obj(self):
        """
        按行读取数据，表单所有数据
        每个用例存储在一个对象中
        :return: 返回一个列表，列表中每个元素为一个用例对象
        """
        # 按行获取数据转换成列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        # 定义一个空列表用来存储所有的用例
        # cases = []
        # for case in rows_data[1:]:
        #     # 创建一个Cases类的对象，用来保存用例数据，
        #     case_obj = Case()
        #     # data用例临时存放用例数据
        #     data = []
        #     # 判断该单元格是否为字符串类型，
        #     for cell in case:
        #         if isinstance(cell.value, str):
        #             data.append(eval(cell.value))
        #         else:
        #             data.append(cell.value)
        #     print(data)
            # 将该条数据放入cases中
            # case_data = list(zip(titles, data))
            # for i in case_data:
            #     setattr(case_obj, i[0], i[1])
            # print(case_obj.case_id, case_obj.data)
            # cases.append(case_obj)

        # return cases

        #     case_data = list(zip(titles, data))
        #     for i in case_data:
        #         setattr(case_obj, i[0], i[1])
        #     print(case_obj.case_id, case_obj.data)
        # self.wb.close()


if __name__ == '__main__':
    list1 = [1, 3]
    r1 = ReadExcel_dict('cases.xlsx', 'Sheet', list1)
    # r2 = ReadExcel_obj('cases.xlsx', 'Sheet', list1)
    # r.read_data_line(list1)
    print('-----------------')
    r1.read_data()


