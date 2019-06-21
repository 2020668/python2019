# -*- coding: utf-8 -*-
"""

=================================
Author: keen
Created on: 2019/5/26

E-mail:keen2020@outlook.com

=================================


"""

import openpyxl


class Case(object):
    # 用这个类来存储用例
    def __init__(self, atrrs):
        """
        初始化用例
        :param atrrs: zip类型--> [（key1,value1),（key2,value2)....]
        """
        for item in atrrs:
            setattr(self, item[0], item[1])


class ReadExcel(object):
    """
    读取Excel数据
    """
    def __init__(self, file_name, sheet_name):
        """
        初始化读取对象
        :param file_name: 文件名，测试用例文件--> str
        :param sheet_name: 表单名--> str
        """
        self.wb = openpyxl.load_workbook(file_name)     # 打开工作簿，传入的指定文件名
        self.sheet = self.wb[sheet_name]    # 选取表单，传入的指定表单
        self.file_name = file_name

    def __del__(self):
        self.wb.close()

    def read_data(self, list1):
        """
        执行读取数据
        :param list1: 指定读取的列[1,2,3,4...]
        :return: 返回一个list，每个元素为一个用例,dict
        """
        max_r = self.sheet.max_row
        cases = []  # 存储所有用例
        titles = []  # 存放标题
        for row in range(1, max_r+1):
            if row != 1:    # 排除表头，获取数据
                case_data = []      # 存储该行的数据
                for column in list1:
                    case_rc = self.sheet.cell(row, column).value   # 获取指定单元格的数据
                    case_data.append(case_rc)
                case = list(zip(titles, case_data))     # 将表头和数据就那些打包，--> 第1次遍历已经添加了表头
                case_obj = Case(case)
                cases.append(case_obj)
            else:   # 获取表头信息
                for column in list1:
                    title = self.sheet.cell(row, column).value
                    titles.append(title)
        return cases

    def write_data(self, row, column, msg):
        self.sheet.cell(row=row, column=column, value=msg)
        self.wb.save(self.file_name)


# if __name__ == '__main__':
#     list1 = [1, 3, 5]
#     r = ReadExcel("cases_columns.xlsx", "Sheet1")
#     r.read_data(list1)      # 测试指定行的测试用例
    # suite = unittest.TestSuite()  # 创建测试集合
    # cases = r.read_data()   # 调用读取Excel方法，接收测试用例数据，对象的list
    # for case in cases:  # 取出单个用例对象
    #     suite.addTest(RegisterTestCase('test_register', case.excepted, case.data))  # 调用用例对象的方法获取用例数据
    #
    # with open('report.html', 'wb') as f:
    #     runner = HTMLTestRunner(
    #         stream=f,
    #         verbosity=2,
    #         title='py18_report',
    #         description='这是我们18期的第一份测试报告',
    #         tester='keen'
    #     )
    #     runner.run(suite)
