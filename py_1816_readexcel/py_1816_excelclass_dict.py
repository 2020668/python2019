# -*- coding: utf-8 -*-
"""

=================================
Author: keen
Created on: 2019/5/26

E-mail:keen2020@outlook.com

=================================


"""


"""
使用字典存储文件的1,3列数据, case_id 和 data
"""

import openpyxl


class ReadExcel_dict(object):
    """
    读取Excel数据
    """

    # 打开工作簿
    def __init__(self, file_name, sheet_name, list):
        """
        初始化读取对象
        :param file_name: 文件名称 --> str
        :param sheet_name:  表单名称 --> str
        :param list1:  指定列 --> list
        """

        self.wb = openpyxl.load_workbook(file_name)     # 打开工作簿，传入的指定文件名
        self.sheet = self.wb[sheet_name]        # 选取表单，传入的指定表单
        self.list = list1

    def read_date(self):
        """
        执行读取数据
        :return:
        """

        rows_data = list(self.sheet.rows)       # 按行获取数据，并转换成列表

        titles = []     # 获取表单的表头信息
        titles.append(rows_data[0][self.list[0]-1].value)   # 获取第1列表头
        titles.append(rows_data[0][self.list[1]-1].value)   # 获取第3列表头

        cases = []
        for i in range(2, self.sheet.max_row + 1):
            data = []
            data.append(self.sheet.cell(row=i, column=self.list[0]).value)      # 读取1列数据
            data.append(eval(self.sheet.cell(row=i, column=self.list[1]).value))        # 读取第3列数据

            # 将表头和内部数据打包，一一对应,将全部用例数据转换成字典
            case_data = dict(list(zip(titles, data)))
            cases.append(case_data)    # 存入列表
        self.wb.close()
        return cases


if __name__ == '__main__':
    list1 = [1, 3]     # 读取指定列的数据
    # 传入测试用例文件名，表单名，指定读取的列，实例化
    case_obj = ReadExcel_dict("cases.xlsx", "Sheet", list1)
    # 调用读取数据方法，返回所有测试用例数据，list
    case_list = case_obj.read_date()
    for i in case_list:    # 打印指定列的测试数据
        print(i['case_id'], i['data'])
