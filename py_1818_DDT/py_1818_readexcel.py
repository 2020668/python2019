# -*- coding: utf-8 -*-
"""

=================================
Author: keen
Created on: 2019/5/26

E-mail:keen2020@outlook.com

=================================


"""

import openpyxl


class Case(object):
    # 用这个类来存储用例
    def __init__(self, attrs):
        """
        初始化用例
        :param atrrs: zip类型--> [（key1,value1),（key2,value2)....]
        """
        for item in attrs:
            setattr(self, item[0], item[1])


class ReadExcel(object):
    """
    读取Excel数据
    """
    def __init__(self, file_name, sheet_name):
        """
        初始化读取对象
        :param file_name: 文件名，测试用例文件--> str
        :param sheet_name: 表单名--> str
        """
        self.wb = openpyxl.load_workbook(file_name)     # 打开工作簿，传入的指定文件名
        self.sheet = self.wb[sheet_name]    # 选取表单，传入的指定表单
        self.file_name = file_name

    def __del__(self):
        self.wb.close()

    def read_data(self, list1):
        """
        执行读取数据
        :param list1: 指定读取的列[1,2,3,4...]
        :return: 返回一个list，每个元素为一个用例,dict
        """
        max_r = self.sheet.max_row
        cases = []  # 存储所有用例
        titles = []  # 存放标题
        for row in range(1, max_r+1):
            if row != 1:    # 排除表头，获取数据
                case_data = []      # 存储该行的数据
                for column in list1:
                    case_rc = self.sheet.cell(row, column).value   # 获取指定单元格的数据
                    case_data.append(case_rc)
                case = list(zip(titles, case_data))     # 将表头和数据就那些打包，--> 第1次遍历已经添加了表头
                case_obj = Case(case)
                cases.append(case_obj)
            else:   # 获取表头信息
                for column in list1:
                    title = self.sheet.cell(row, column).value
                    titles.append(title)
        return cases

    def write_data(self, row, column, msg):
        self.sheet.cell(row=row, column=column, value=msg)
        self.wb.save(self.file_name)
