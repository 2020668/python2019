# -*- coding: utf-8 -*-
"""

=================================
Author: keen
Created on: 2019/5/26

E-mail:keen2020@outlook.com

=================================


"""

import openpyxl


# excel 读取数据

# 打开一个工作簿
wb = openpyxl.load_workbook('cases.xlsx')

# 选定表单
sh = wb['Sheet']

# 读取指定单元格数据
ce = sh.cell(row=2, column=2)
# 取出值
# print(ce.value)

# 按行读取数据,生成一个元组
rows_data = sh.rows
# print(list(rows_data))

# 指定单元格写入数据
sh.cell(row=1, column=5, value='result')
# 写完数据要保存，指定文件名字

# 获取总行数，列数
print(sh.max_row)
print(sh.max_column)

# 创建表单，指定第六个位置
# wb.create_sheet('Sheetcase', 5)
# wb.create_sheet('Sheet2')
# 删除表单
# del wb['Sheet2']
# wb.remove(wb['Sheetcase'])
# 显示所有的表单名字
print(wb.sheetnames)

# 循环读取第一列数据,（第一列，所有行）
for i in range(1, sh.max_row+1):
    data = sh.cell(row=i, column=1).value
    print(data)

# 读取整个表单第2列和第3列的数据，每行的数据放到一个元组中，所有行数据放到列表中
cases = []
for i in range(2, sh.max_row+1):
    data2 = eval(sh.cell(row=i, column=2).value)
    data3 = eval(sh.cell(row=i, column=3).value)
    tu = (data2, data3)
    cases.append(tu)

print(cases)
wb.save('cases.xlsx')
wb.close()
